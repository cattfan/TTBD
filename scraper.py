import asyncio
import random
import re
import os
from playwright.async_api import async_playwright
import openpyxl

USER_AGENTS = [
    "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 12; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPad; CPU OS 16_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.5 Mobile/15E148 Safari/604.1"
]

async def scrape_single_link(page, url):
    data = {"Views": "0", "Likes": "0", "Comments": "0", "Saves": "0", "Shares": "0"}
    try:
        await page.goto(url, wait_until="networkidle", timeout=60000)
        await asyncio.sleep(random.uniform(2, 4))
        content = await page.content()
        patterns = {
            "Views": r'"playCount":(\d+)',
            "Likes": r'"diggCount":(\d+)',
            "Comments": r'"commentCount":(\d+)',
            "Saves": r'"collectCount":(\d+)',
            "Shares": r'"shareCount":(\d+)'
        }
        for key, pattern in patterns.items():
            match = re.search(pattern, content)
            if match: data[key] = match.group(1)
        return data, "Success"
    except Exception as e:
        return data, f"Error: {str(e)}"

async def run_scraper(file_path, websocket_manager=None):
    if not os.path.exists(file_path):
        if websocket_manager: await websocket_manager.broadcast_log(f"Lỗi: Không tìm thấy file {file_path}")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Định nghĩa bản đồ cột dựa trên Header thực tế trong file của bạn
    column_map = {
        "LƯỢT XEM": None,
        "TIM": None,
        "BÌNH LUẬN": None,
        "LƯỢT LƯU": None,
        "CHIA SẺ": None,
        "URL": None
    }

    # Quét hàng đầu tiên để tìm vị trí các cột
    for cell in sheet[1]:
        header = str(cell.value).strip().upper() if cell.value else ""
        if "URL" in header: column_map["URL"] = cell.column
        if "LƯỢT XEM" in header: column_map["LƯỢT XEM"] = cell.column
        if "TIM" in header: column_map["TIM"] = cell.column
        if "BÌNH LUẬN" in header: column_map["BÌNH LUẬN"] = cell.column
        if "LƯỢT LƯU" in header: column_map["LƯỢT LƯU"] = cell.column
        if "CHIA SẺ" in header: column_map["CHIA SẺ"] = cell.column

    # Nếu không tìm thấy bằng tên, gán cứng theo ảnh (D, E, F, G, H là 4, 5, 6, 7, 8)
    if not column_map["LƯỢT XEM"]: column_map["LƯỢT XEM"] = 4
    if not column_map["TIM"]: column_map["TIM"] = 5
    if not column_map["BÌNH LUẬN"]: column_map["BÌNH LUẬN"] = 6
    if not column_map["LƯỢT LƯU"]: column_map["LƯỢT LƯU"] = 7
    if not column_map["CHIA SẺ"]: column_map["CHIA SẺ"] = 8
    # Tìm cột URL nếu chưa xác định được qua header
    if not column_map["URL"]:
        # Quét hàng 2 để tìm cột nào chứa link tiktok
        for col in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=2, column=col).value or "")
            if "tiktok.com" in val:
                column_map["URL"] = col
                break
        # Nếu vẫn không thấy, mặc định là cột 2 (B)
        if not column_map["URL"]: column_map["URL"] = 2

    rows_to_process = []
    for row in range(2, sheet.max_row + 1):
        url_cell = sheet.cell(row=row, column=column_map["URL"]).value
        url = str(url_cell).strip() if url_cell else ""
        if url and "tiktok.com" in url:
            rows_to_process.append((row, url))

    total = len(rows_to_process)
    if websocket_manager: 
        await websocket_manager.broadcast_status({"total": total, "processed": 0, "success": 0, "error": 0})
        await websocket_manager.broadcast_log(f"Bắt đầu xử lý {total} links (Cột {column_map['URL']})...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        success_count = 0
        error_count = 0
        
        for i, (row_idx, url) in enumerate(rows_to_process):
            ua = random.choice(USER_AGENTS)
            context = await browser.new_context(user_agent=ua, viewport={'width': 390, 'height': 844})
            page = await context.new_page()
            
            data, status = await scrape_single_link(page, url)
            
            if "Success" in status:
                success_count += 1
                sheet.cell(row=row_idx, column=column_map["LƯỢT XEM"]).value = int(data["Views"])
                sheet.cell(row=row_idx, column=column_map["TIM"]).value = int(data["Likes"])
                sheet.cell(row=row_idx, column=column_map["BÌNH LUẬN"]).value = int(data["Comments"])
                sheet.cell(row=row_idx, column=column_map["LƯỢT LƯU"]).value = int(data["Saves"])
                sheet.cell(row=row_idx, column=column_map["CHIA SẺ"]).value = int(data["Shares"])
            else:
                error_count += 1
            
            try:
                wb.save(file_path)
            except PermissionError:
                if websocket_manager:
                    await websocket_manager.broadcast_log("CẢNH BÁO: File Excel đang mở, không thể ghi đè! Vui lòng đóng file.")
            except Exception as e:
                if websocket_manager:
                    await websocket_manager.broadcast_log(f"CẢNH BÁO: Lỗi lưu file ({str(e)})")
            
            if websocket_manager:
                await websocket_manager.broadcast_data({
                    "id": i + 1, "url": url,
                    "views": data["Views"], "likes": data["Likes"],
                    "comments": data["Comments"], "saves": data["Saves"],
                    "shares": data["Shares"], "status": status
                })
                await websocket_manager.broadcast_status({
                    "total": total, "processed": i + 1, "success": success_count, "error": error_count
                })
                await websocket_manager.broadcast_log(f"Xong link {i+1}: {status}")

            await context.close()
            await asyncio.sleep(random.uniform(2, 5))

        await browser.close()
    if websocket_manager: await websocket_manager.broadcast_log("HOÀN THÀNH: Đã cập nhật toàn bộ dữ liệu vào Excel.")
